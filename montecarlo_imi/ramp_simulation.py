from __future__ import annotations

import math
from dataclasses import asdict, dataclass, field
from typing import Dict, Literal

import numpy as np
import pandas as pd


ScenarioName = Literal["prudent", "central", "ambitieux"]


@dataclass(frozen=True)
class TriangularPrior:
    low: float
    mode: float
    high: float

    def sample(self, rng: np.random.Generator) -> float:
        return float(rng.triangular(float(self.low), float(self.mode), float(self.high)))


@dataclass
class RampConfig:
    months: int = 60
    n_sims: int = 2000
    seed: int = 42

    # --- Données marché (imports, t/an) ---
    imports_tpy_by_year: dict[int, float] = field(
        default_factory=lambda: {
            2019: 16657.0,
            2020: 22582.0,
            2021: 22711.0,
            2022: 23049.0,
            2023: 21314.0,
            2024: 23853.0,
        }
    )
    baseline_year: int = 2024  # D0 = imports[baseline_year]

    growth_prudent: float = 0.03
    growth_ambitieux: float = 0.08
    central_growth_mode: Literal["cagr", "override"] = "cagr"
    growth_central_override: float = 0.05

    sigma_y_mode: Literal["cv", "override"] = "cv"
    sigma_y_override: float = 0.10
    market_noise: bool = True

    # --- Part de marché atteignable s(t) ---
    s_max: float = 0.60
    t0_s: float = 36.0
    k_s: float = 0.12

    # --- Production / capacité ---
    Q_max: float = 15000.0  # capacité nominale (t/an)
    u_target: float = 0.85  # utilisation durable à maturité (0..1)
    t0_cap: float = 22.0  # mois d'inflexion (capacité)
    k_cap: float = 0.22  # vitesse de montée (capacité)

    downtime_mean: float = 0.03  # fraction du mois perdue
    downtime_sd: float = 0.02

    # --- Marché / demande (clients) ---
    lam0: float = 0.10  # nouveaux clients / mois à t=0
    lam_max: float = 1.20  # nouveaux clients / mois à maturité
    t0_lam: float = 18.0  # inflexion acquisition
    k_lam: float = 0.18  # vitesse acquisition

    onboard_mean: float = 4.0  # mois
    onboard_sd: float = 1.5  # mois

    mean_target_tpm: float = 18.0  # demande moyenne cible par client (t/mois)
    sigma_logn: float = 0.6  # dispersion (log-space)

    cust_growth_mean: float = 0.015  # croissance mensuelle de demande/client
    cust_growth_sd: float = 0.010
    churn_monthly: float = 0.01  # probabilité mensuelle de perdre un client

    scenario: ScenarioName = "central"
    demo_seed: int = 123

    threshold_1: float = 11000.0
    threshold_2: float = 15000.0

    def as_dict(self) -> Dict[str, object]:
        return asdict(self)


def ocp_new_line_presets() -> dict[str, dict[str, object]]:
    """
    Scénarios "réalistes" pour un nouvel entrant qui installe sa première ligne.

    Remarque: ce sont des presets métier (hypothèses) destinés à donner un point de départ
    cohérent. L'utilisateur est censé les ajuster selon sa réalité (engineering, sales, marché).
    """

    base = {
        # Horizon plus long pour une 1ère ligne + adoption
        "months": 96,
        "scenario": "central",
        # Part de marché atteignable (nouvel entrant)
        "s_max": 0.30,
        "t0_s": 60.0,
        "k_s": 0.10,  # ~10->90% en ~44 mois
        # Capacité / ramp-up première ligne
        "Q_max": 18_000.0,  # nameplate pour viser ~15k t/an à 85%
        "u_target": 0.85,
        "t0_cap": 24.0,
        "k_cap": 0.15,  # ~10->90% en ~29 mois
        # Downtime plus élevé au démarrage
        "downtime_mean": 0.06,
        "downtime_sd": 0.03,
        # Marché: croissance centrale modérée
        "growth_prudent": 0.02,
        "growth_ambitieux": 0.07,
        "central_growth_mode": "override",
        "growth_central_override": 0.04,
        # Acquisition clients: montée progressive
        "lam0": 0.05,
        "lam_max": 0.80,
        "t0_lam": 30.0,
        "k_lam": 0.12,
        "onboard_mean": 6.0,
        "onboard_sd": 2.0,
        "mean_target_tpm": 15.0,
        "sigma_logn": 0.70,
        "cust_growth_mean": 0.010,
        "cust_growth_sd": 0.010,
        "churn_monthly": 0.015,
    }

    prudent = {
        "months": 120,
        "scenario": "prudent",
        "s_max": 0.25,
        "t0_s": 72.0,
        "k_s": 0.08,  # ~10->90% en ~55 mois
        "Q_max": 17_000.0,
        "u_target": 0.82,
        "t0_cap": 30.0,
        "k_cap": 0.10,  # ~10->90% en ~44 mois
        "downtime_mean": 0.08,
        "downtime_sd": 0.04,
        "growth_prudent": 0.015,
        "growth_ambitieux": 0.06,
        "central_growth_mode": "override",
        "growth_central_override": 0.03,
        "lam0": 0.03,
        "lam_max": 0.60,
        "t0_lam": 36.0,
        "k_lam": 0.10,
        "onboard_mean": 7.0,
        "onboard_sd": 2.5,
        "mean_target_tpm": 12.0,
        "sigma_logn": 0.75,
        "cust_growth_mean": 0.008,
        "cust_growth_sd": 0.012,
        "churn_monthly": 0.020,
    }

    return {
        "OCP — nouvelle ligne (base)": base,
        "OCP — nouvelle ligne (prudent)": prudent,
    }


def logistic(x: np.ndarray | float, k: float, x0: float):
    return 1.0 / (1.0 + np.exp(-k * (x - x0)))


def annualize(x_tpm: np.ndarray) -> np.ndarray:
    """Convertit t/mois -> t/an."""
    return x_tpm * 12.0


def imports_series(cfg: RampConfig) -> pd.Series:
    s = pd.Series(cfg.imports_tpy_by_year, name="Imports_tons").sort_index()
    s.index.name = "Year"
    return s.astype(float)


@dataclass(frozen=True)
class MarketStats:
    mean_imports: float
    std_imports: float
    cv_imports: float
    cagr: float
    D0_tpy: float
    sigma_y: float
    sigma_m: float


def market_summary(imports: pd.Series) -> tuple[pd.DataFrame, float, float, float, float]:
    mean_imports = float(imports.mean())
    std_imports = float(imports.std(ddof=1))
    cv_imports = float(std_imports / mean_imports) if mean_imports else 0.0

    years = list(imports.index)
    start_year = int(min(years))
    end_year = int(max(years))
    n_years = max(1, end_year - start_year)
    cagr = float((imports.loc[end_year] / imports.loc[start_year]) ** (1 / n_years) - 1)

    df = pd.DataFrame(
        {
            "Mean (t/an)": [mean_imports],
            "Std (t/an)": [std_imports],
            "CV": [cv_imports],
            f"CAGR {start_year}-{end_year}": [cagr],
        }
    )
    return df, mean_imports, std_imports, cv_imports, cagr


def _build_growth_scenarios(cfg: RampConfig, cagr: float) -> dict[ScenarioName, float]:
    central = cagr if cfg.central_growth_mode == "cagr" else float(cfg.growth_central_override)
    return {
        "prudent": float(cfg.growth_prudent),
        "central": float(central),
        "ambitieux": float(cfg.growth_ambitieux),
    }


def _sigma_y(cfg: RampConfig, cv_imports: float) -> float:
    return float(cv_imports) if cfg.sigma_y_mode == "cv" else float(cfg.sigma_y_override)


def market_demand_path(
    *,
    months: int,
    D0_tpy: float,
    growth_annual: float,
    sigma_m: float,
    noise: bool,
    rng_local: np.random.Generator,
) -> np.ndarray:
    g_m = (1 + float(growth_annual)) ** (1 / 12) - 1
    t = np.arange(months + 1)
    D = np.zeros_like(t, dtype=float)
    D[0] = float(D0_tpy)
    for m in range(1, months + 1):
        eps = float(rng_local.normal(0.0, sigma_m)) if noise else 0.0
        D[m] = max(0.0, D[m - 1] * (1 + g_m) * (1 + eps))
    return D


def _share_series(cfg: RampConfig, t: np.ndarray) -> np.ndarray:
    return float(cfg.s_max) * logistic(t, float(cfg.k_s), float(cfg.t0_s))


def _capacity_series_tpy(cfg: RampConfig, t: np.ndarray) -> np.ndarray:
    Q_eff_max = float(cfg.Q_max) * float(cfg.u_target)
    return Q_eff_max * logistic(t, float(cfg.k_cap), float(cfg.t0_cap))


def _clamp01(x: float) -> float:
    return float(min(1.0, max(0.0, x)))


def sample_cfg_with_priors(
    base_cfg: RampConfig, priors: dict[str, TriangularPrior], rng: np.random.Generator
) -> RampConfig:
    """
    Échantillonne un scénario de paramètres (incertitude épistémique) autour d'un RampConfig de base.

    - Les bornes/contraintes sont respectées (ex: variables dans [0,1], non-négativité).
    - Si lam_max < lam0 après tirage, on corrige pour garder lam_max >= lam0.

    Remarque: on ne modifie pas les champs non présents dans `priors`.
    """
    cfg_dict = base_cfg.as_dict()
    for name, prior in priors.items():
        if name not in cfg_dict:
            continue
        cfg_dict[name] = prior.sample(rng)

    # Contraintes
    for p in ["s_max", "u_target", "downtime_mean", "downtime_sd", "churn_monthly"]:
        if p in cfg_dict:
            cfg_dict[p] = _clamp01(float(cfg_dict[p]))

    for p in [
        "Q_max",
        "k_s",
        "k_cap",
        "k_lam",
        "lam0",
        "lam_max",
        "onboard_mean",
        "onboard_sd",
        "mean_target_tpm",
        "sigma_logn",
        "cust_growth_sd",
    ]:
        if p in cfg_dict:
            cfg_dict[p] = float(max(0.0, float(cfg_dict[p])))

    for p in ["t0_s", "t0_cap", "t0_lam"]:
        if p in cfg_dict:
            cfg_dict[p] = float(max(0.0, float(cfg_dict[p])))

    if "lam0" in cfg_dict and "lam_max" in cfg_dict:
        lam0 = float(cfg_dict["lam0"])
        lam_max = float(cfg_dict["lam_max"])
        if lam_max < lam0:
            cfg_dict["lam_max"] = lam0

    # sigma_logn peut être 0 (pas de dispersion) mais doit rester >= 0
    if "sigma_logn" in cfg_dict:
        cfg_dict["sigma_logn"] = float(max(0.0, float(cfg_dict["sigma_logn"])))

    # mean_target_tpm doit être > 0 pour log(); on impose un plancher
    if "mean_target_tpm" in cfg_dict:
        cfg_dict["mean_target_tpm"] = float(max(1e-6, float(cfg_dict["mean_target_tpm"])))

    return RampConfig(**cfg_dict)


def simulate_one_path(
    *,
    cfg: RampConfig,
    growth_annual: float,
    rng_local: np.random.Generator,
    t: np.ndarray,
    sigma_m: float,
    D0_tpy: float,
    s_t: np.ndarray,
    Q_t: np.ndarray,
    P_tpm_base: np.ndarray,
    mu_logn: float,
) -> dict[str, np.ndarray]:
    Dm = market_demand_path(
        months=cfg.months,
        D0_tpy=D0_tpy,
        growth_annual=growth_annual,
        sigma_m=sigma_m,
        noise=bool(cfg.market_noise),
        rng_local=rng_local,
    )
    Daddr = s_t * Dm

    customers: list[dict] = []
    demand_clients_tpm = np.zeros_like(t, dtype=float)
    sales_tpm = np.zeros_like(t, dtype=float)

    for m in range(1, cfg.months + 1):
        lam_m = float(cfg.lam0) + (float(cfg.lam_max) - float(cfg.lam0)) * logistic(
            m, float(cfg.k_lam), float(cfg.t0_lam)
        )
        n_new = int(rng_local.poisson(lam_m))
        for _ in range(n_new):
            delay = max(0, int(round(float(rng_local.normal(float(cfg.onboard_mean), float(cfg.onboard_sd))))))
            start = min(cfg.months, m + delay)
            base = float(rng_local.lognormal(mu_logn, float(cfg.sigma_logn)))
            customers.append({"start": start, "demand": base, "active": True})

        total_demand = 0.0
        for c in customers:
            if (not c["active"]) or (m < c["start"]):
                continue
            if float(rng_local.random()) < float(cfg.churn_monthly):
                c["active"] = False
                continue
            g = float(rng_local.normal(float(cfg.cust_growth_mean), float(cfg.cust_growth_sd)))
            c["demand"] *= max(0.95, (1.0 + g))
            total_demand += float(c["demand"])

        demand_clients_tpm[m] = total_demand

        dt = max(0.0, float(rng_local.normal(float(cfg.downtime_mean), float(cfg.downtime_sd))))
        prod_tpm = float(P_tpm_base[m]) * max(0.0, 1.0 - dt)

        Daddr_tpm = float(Daddr[m]) / 12.0
        sales_tpm[m] = min(prod_tpm, total_demand, Daddr_tpm)

    return {
        "Capacity_tpy": Q_t,
        "Dm_tpy": Dm,
        "Daddr_tpy": Daddr,
        "DemandClients_tpy": demand_clients_tpm * 12.0,
        "Sales_tpy": sales_tpm * 12.0,
    }


def quantiles(arr_2d: np.ndarray) -> dict[str, np.ndarray]:
    return {
        "P10": np.quantile(arr_2d, 0.10, axis=0),
        "P50": np.quantile(arr_2d, 0.50, axis=0),
        "P90": np.quantile(arr_2d, 0.90, axis=0),
        "Mean": np.mean(arr_2d, axis=0),
    }


def first_cross(series: np.ndarray, threshold: float) -> int | None:
    idx = np.where(series >= threshold)[0]
    return int(idx[0]) if len(idx) else None


@dataclass
class RampResults:
    cfg: RampConfig
    t: np.ndarray
    imports: pd.Series
    summary_market: pd.DataFrame
    growth_scenarios: dict[ScenarioName, float]
    stats: MarketStats
    s_t: np.ndarray
    Q_t: np.ndarray
    demo: dict[str, np.ndarray]
    mc: dict[str, np.ndarray]
    q: dict[str, dict[str, np.ndarray]]
    summary_df: pd.DataFrame
    crossings_df: pd.DataFrame


def run_ramp_monte_carlo(cfg: RampConfig, priors: dict[str, TriangularPrior] | None = None) -> RampResults:
    t = np.arange(cfg.months + 1)

    imports = imports_series(cfg)
    summary_market_df, mean_imports, std_imports, cv_imports, cagr = market_summary(imports)
    growth_scenarios = _build_growth_scenarios(cfg, cagr)

    if cfg.baseline_year in imports.index:
        D0_tpy = float(imports.loc[int(cfg.baseline_year)])
    else:
        D0_tpy = float(imports.iloc[-1])

    sigma_y = _sigma_y(cfg, cv_imports)
    sigma_m = float(sigma_y) / math.sqrt(12)

    stats = MarketStats(
        mean_imports=mean_imports,
        std_imports=std_imports,
        cv_imports=cv_imports,
        cagr=cagr,
        D0_tpy=D0_tpy,
        sigma_y=float(sigma_y),
        sigma_m=float(sigma_m),
    )

    rng = np.random.default_rng(int(cfg.seed))

    keys = ["Capacity_tpy", "Dm_tpy", "Daddr_tpy", "DemandClients_tpy", "Sales_tpy"]
    mc = {k: np.zeros((int(cfg.n_sims), int(cfg.months) + 1), dtype=float) for k in keys}

    g = float(growth_scenarios[cfg.scenario])
    for i in range(int(cfg.n_sims)):
        cfg_i = sample_cfg_with_priors(cfg, priors, rng) if priors else cfg
        s_t_i = _share_series(cfg_i, t)
        Q_t_i = _capacity_series_tpy(cfg_i, t)
        P_tpm_base_i = Q_t_i / 12.0

        mu_logn_i = math.log(float(cfg_i.mean_target_tpm)) - 0.5 * float(cfg_i.sigma_logn) ** 2

        res = simulate_one_path(
            cfg=cfg_i,
            growth_annual=g,
            rng_local=rng,
            t=t,
            sigma_m=sigma_m,
            D0_tpy=D0_tpy,
            s_t=s_t_i,
            Q_t=Q_t_i,
            P_tpm_base=P_tpm_base_i,
            mu_logn=mu_logn_i,
        )
        for k in keys:
            mc[k][i, :] = res[k]

    q = {k: quantiles(mc[k]) for k in keys}

    summary_df = pd.DataFrame(
        {
            "Month": t,
            "Capacity_P50_tpy": q["Capacity_tpy"]["P50"],
            "Market_P50_tpy": q["Dm_tpy"]["P50"],
            "Addressable_P50_tpy": q["Daddr_tpy"]["P50"],
            "DemandClients_P50_tpy": q["DemandClients_tpy"]["P50"],
            "Sales_P50_tpy": q["Sales_tpy"]["P50"],
            "Sales_P10_tpy": q["Sales_tpy"]["P10"],
            "Sales_P90_tpy": q["Sales_tpy"]["P90"],
            "Sales_Mean_tpy": q["Sales_tpy"]["Mean"],
        }
    )

    m1 = first_cross(q["Sales_tpy"]["P50"], float(cfg.threshold_1))
    m2 = first_cross(q["Sales_tpy"]["P50"], float(cfg.threshold_2))
    crossings_df = pd.DataFrame(
        {
            "Seuil (t/an)": [float(cfg.threshold_1), float(cfg.threshold_2)],
            "Mois atteint (P50 ventes)": [m1, m2],
        }
    )

    # Courbes "base" (déterministes) issues du cfg de base (utile pour visualiser la forme des logistiques)
    s_t = _share_series(cfg, t)
    Q_t = _capacity_series_tpy(cfg, t)
    P_tpm_base = Q_t / 12.0
    mu_logn = math.log(float(cfg.mean_target_tpm)) - 0.5 * float(cfg.sigma_logn) ** 2

    demo_rng = np.random.default_rng(int(cfg.demo_seed))
    demo = simulate_one_path(
        cfg=cfg,
        growth_annual=g,
        rng_local=demo_rng,
        t=t,
        sigma_m=sigma_m,
        D0_tpy=D0_tpy,
        s_t=s_t,
        Q_t=Q_t,
        P_tpm_base=P_tpm_base,
        mu_logn=mu_logn,
    )

    return RampResults(
        cfg=cfg,
        t=t,
        imports=imports,
        summary_market=summary_market_df,
        growth_scenarios=growth_scenarios,
        stats=stats,
        s_t=s_t,
        Q_t=Q_t,
        demo=demo,
        mc=mc,
        q=q,
        summary_df=summary_df,
        crossings_df=crossings_df,
    )


def build_excel_bytes(results: RampResults) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils.dataframe import dataframe_to_rows

    cfg = results.cfg

    wb = Workbook()

    wsP = wb.active
    wsP.title = "Parametres"
    wsP.append(["Paramètre", "Valeur"])
    wsP["A1"].font = Font(bold=True)
    wsP["B1"].font = Font(bold=True)

    flat_params: list[tuple[str, object]] = []
    for k, v in cfg.as_dict().items():
        if k == "imports_tpy_by_year":
            continue
        flat_params.append((k, v))
    flat_params.sort(key=lambda kv: kv[0])

    for k, v in flat_params:
        wsP.append([k, v])
    wsP.column_dimensions["A"].width = 34
    wsP.column_dimensions["B"].width = 22

    wsI = wb.create_sheet("Imports")
    wsI.append(["Year", "Imports_tons"])
    wsI["A1"].font = Font(bold=True)
    wsI["B1"].font = Font(bold=True)
    for year, val in results.imports.items():
        wsI.append([int(year), float(val)])
    wsI.append([])
    wsI.append(["Mean (t/an)", float(results.stats.mean_imports)])
    wsI.append(["Std (t/an)", float(results.stats.std_imports)])
    wsI.append(["CV", float(results.stats.cv_imports)])
    wsI.append(["CAGR", float(results.stats.cagr)])
    wsI.column_dimensions["A"].width = 18
    wsI.column_dimensions["B"].width = 18

    ws = wb.create_sheet("Resultats")
    for r in dataframe_to_rows(results.summary_df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 10
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 22

    chart = LineChart()
    chart.title = "Monte Carlo — P50 + bande ventes"
    chart.y_axis.title = "t/an"
    chart.x_axis.title = "mois"
    data = Reference(ws, min_col=2, max_col=8, min_row=1, max_row=int(cfg.months) + 2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=int(cfg.months) + 2)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 28
    ws.add_chart(chart, "K2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
